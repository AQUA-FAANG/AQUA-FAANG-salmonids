"""
gimme maelstrom – Auto-annotate motif table with factors + logos
================================================================
Reads the top-motifs CSV/Excel from your R output, fetches factor names
and logo images directly from the gimmemotifs database, and produces:
  1. An HTML report  (logos embedded as base64 – opens in any browser)
  2. An Excel file   (logos inserted as images via openpyxl)

Requirements:
    pip install gimmemotifs openpyxl pandas Pillow
"""

import os
import base64
import pandas as pd
from PIL import Image
from gimmemotifs.motif import read_motifs
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==============================================================================
# 1. CONFIG  – update these paths
# ==============================================================================

INPUT_FILE   = "maelstrom_conservation_top60_per_category.xlsx" ### loop this between different iterations of the analysis
SHEET_NAME   = "Sheet1"
LOGO_DIR     = "all_logos"
OUTPUT_HTML  = "maelstrom_conservation_top60_per_category_logos.html"
OUTPUT_EXCEL = "maelstrom_conservation_top60_per_category_logos.xlsx"

# Logo generation + display size
LOGO_WIDTH_PX       = 200
LOGO_HEIGHT_PX      = 60

# Logo display size in Excel and HTML
LOGO_DISPLAY_WIDTH  = 200
LOGO_DISPLAY_HEIGHT = 60

# JPEG quality (1–95): 70 is a good balance; drop to 50 if files still too large
LOGO_JPEG_QUALITY   = 70

# Excel row height (pts)
LOGO_ROW_HEIGHT     = 70

os.makedirs(LOGO_DIR, exist_ok=True)

# ==============================================================================
# 2. LOAD GIMMEMOTIFS DATABASE
# ==============================================================================

print("Loading gimmemotifs database …")
all_motifs = read_motifs()
motif_dict = {m.id: m for m in all_motifs}
print(f"  Loaded {len(motif_dict)} motifs")

# ==============================================================================
# 3. LOAD YOUR TOP-MOTIF TABLE
# ==============================================================================

df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
print(f"Input table: {df.shape[0]} rows × {df.shape[1]} cols")
print(f"Columns: {list(df.columns)}")

# Rename first column to 'motif' if needed
df.columns = ["motif"] + list(df.columns[1:])

# ==============================================================================
# 4. ANNOTATE: factors + logo path
# ==============================================================================

def get_factors(motif_id):
    m = motif_dict.get(motif_id)
    if m is None:
        return "not found"
    factors = getattr(m, "factors", {})
    if isinstance(factors, dict):
        all_factors = []
        for key in ["direct", "indirect", "predicted"]:
            all_factors.extend(factors.get(key, []))
        seen   = set()
        unique = [f for f in all_factors if not (f in seen or seen.add(f))]
        return ", ".join(unique) if unique else motif_id
    if isinstance(factors, list):
        return ", ".join(factors) if factors else motif_id
    return str(factors) if factors else motif_id


def save_logo(motif_id, out_dir):
    """Save compressed JPEG motif logo; return path or None on failure."""
    m = motif_dict.get(motif_id)
    if m is None:
        return None

    path = os.path.join(out_dir, f"{motif_id.replace('.', '_')}.jpg")
    if not os.path.exists(path):
        try:
            # gimmemotifs only writes PNG — save to temp file first
            tmp_path = path.replace(".jpg", "_tmp.png")
            m.plot_logo(fname=tmp_path, title=False)

            # Convert PNG → JPEG with compression
            img = Image.open(tmp_path).convert("RGB")   # JPEG requires RGB not RGBA
            img = img.resize((LOGO_WIDTH_PX, LOGO_HEIGHT_PX), Image.LANCZOS)
            img.save(path, format="JPEG", quality=LOGO_JPEG_QUALITY, optimize=True)

            # Remove temp PNG
            os.remove(tmp_path)

        except Exception as e:
            print(f"  Warning: could not plot logo for {motif_id}: {e}")
            return None
    return path


print("Fetching factors and generating logos …")
df["factors"]   = df["motif"].apply(get_factors)
df["logo_path"] = df["motif"].apply(lambda mid: save_logo(mid, LOGO_DIR))
print(f"  Logos generated: {df['logo_path'].notna().sum()} / {len(df)}")

# ==============================================================================
# 5. HTML REPORT  (logos as embedded base64)
# ==============================================================================

def img_to_base64(path):
    if path is None or not os.path.exists(path):
        return ""
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

# Identify z-score columns
meta_cols  = ["motif", "top_category", "rank"]
score_cols = [c for c in df.columns
              if c not in meta_cols + ["factors", "logo_path"]]

def zscore_bg(val):
    """Red–white–blue background for z-score cells."""
    try:
        v     = float(val)
        max_v = 4.0
        ratio = max(min(v / max_v, 1.0), -1.0)
        if ratio >= 0:
            r = int(214 + (255 - 214) * (1 - ratio))
            g = int(39  * (1 - ratio))
            b = int(39  * (1 - ratio))
        else:
            r = int(33  * (1 + ratio))
            g = int(102 * (1 + ratio))
            b = int(178 + (255 - 178) * (-ratio))
        return (f"background-color: rgb({r},{g},{b}); "
                f"color: {'white' if abs(ratio) > 0.5 else 'black'}")
    except (ValueError, TypeError):
        return ""

rows_html = []
for _, row in df.iterrows():
    cells = []
    for col in meta_cols:
        cells.append(f"<td>{row.get(col, '')}</td>")
    for col in score_cols:
        val   = row.get(col, "")
        style = zscore_bg(val)
        try:
            display = f"{float(val):.3f}"
        except (ValueError, TypeError):
            display = str(val)
        cells.append(f'<td style="{style}">{display}</td>')
    cells.append(f"<td>{row.get('factors', '')}</td>")
    b64     = img_to_base64(row.get("logo_path"))
    img_tag = (f'<img src="data:image/jpeg;base64,{b64}" '
               f'height="{LOGO_DISPLAY_HEIGHT}"/>' if b64 else "—")
    cells.append(f"<td>{img_tag}</td>")
    rows_html.append("<tr>" + "".join(cells) + "</tr>")

header_cols = meta_cols + score_cols + ["factors", "logo"]
header_html = "".join(f"<th>{c}</th>" for c in header_cols)

html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<title>gimme maelstrom – Top Motifs</title>
<style>
  body  {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th    {{ background: #2c3e50; color: white; padding: 6px 8px;
           position: sticky; top: 0; text-align: center; }}
  td    {{ border: 1px solid #ddd; padding: 4px 6px;
           vertical-align: middle; text-align: center; }}
  tr:nth-child(even) {{ background: #f9f9f9; }}
  tr:hover           {{ background: #eaf4fb; }}
  img   {{ display: block; margin: auto; }}
</style>
</head>
<body>
<h2>gimme maelstrom – Top Enriched Motifs</h2>
<table>
<thead><tr>{header_html}</tr></thead>
<tbody>{"".join(rows_html)}</tbody>
</table>
</body>
</html>"""

with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
    f.write(html)
print(f"\n✓ HTML report saved: {OUTPUT_HTML}")

# ==============================================================================
# 6. EXCEL WITH EMBEDDED LOGO IMAGES
# ==============================================================================

print("Building Excel workbook …")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Top motifs"

# Header style
hdr_fill = PatternFill("solid", fgColor="2C3E50")
hdr_font = Font(bold=True, color="FFFFFF", size=10)
thin     = Side(style="thin", color="CCCCCC")
border   = Border(left=thin, right=thin, top=thin, bottom=thin)
center   = Alignment(horizontal="center", vertical="center", wrap_text=True)

xl_cols = meta_cols + score_cols + ["factors", "logo"]
for ci, col in enumerate(xl_cols, 1):
    cell           = ws.cell(row=1, column=ci, value=col)
    cell.fill      = hdr_fill
    cell.font      = hdr_font
    cell.border    = border
    cell.alignment = center

# Column widths
col_widths = {"motif": 30, "top_category": 12, "rank": 6,
              "factors": 45, "logo": 28}
for ci, col in enumerate(xl_cols, 1):
    ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 9)


def excel_fill(val):
    """Red–white–blue cell fill for z-score values."""
    try:
        v     = float(val)
        ratio = max(min(v / 4.0, 1.0), -1.0)
        if ratio >= 0:
            r = int(255 - (255 - 214) * ratio)
            g = int(255 - (255 - 39)  * ratio)
            b = int(255 - (255 - 39)  * ratio)
        else:
            r = int(255 - (255 - 33)  * (-ratio))
            g = int(255 - (255 - 102) * (-ratio))
            b = int(255 - (255 - 178) * (-ratio))
        return PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
    except (ValueError, TypeError):
        return None


logo_col_idx = xl_cols.index("logo") + 1

for ri, (_, row) in enumerate(df.iterrows(), 2):
    ws.row_dimensions[ri].height = LOGO_ROW_HEIGHT

    for ci, col in enumerate(xl_cols, 1):
        if col == "logo":
            continue
        val            = row.get(col, "")
        cell           = ws.cell(row=ri, column=ci, value=val)
        cell.border    = border
        cell.alignment = center
        cell.font      = Font(size=9)

        if col in score_cols:
            fill = excel_fill(val)
            if fill:
                cell.fill  = fill
                font_color = "FFFFFF" if abs(float(val) / 4.0) > 0.5 else "000000"
                cell.font  = Font(size=9, color=font_color)
            try:
                cell.value = round(float(val), 4)
            except (ValueError, TypeError):
                pass

    # Embed compressed JPEG logo
    logo_path = row.get("logo_path")
    if logo_path and os.path.exists(logo_path):
        try:
            img        = XLImage(logo_path)
            img.width  = LOGO_DISPLAY_WIDTH
            img.height = LOGO_DISPLAY_HEIGHT
            ws.add_image(img, f"{get_column_letter(logo_col_idx)}{ri}")
        except Exception as e:
            ws.cell(row=ri, column=logo_col_idx, value="(logo error)")
            print(f"  Warning: could not embed logo for {row['motif']}: {e}")

wb.save(OUTPUT_EXCEL)
print(f"✓ Excel file saved:  {OUTPUT_EXCEL}")

print("\n=== All done! ===")
print(f"  Open {OUTPUT_HTML} in your browser for the quickest view.")
print(f"  Open {OUTPUT_EXCEL} in Excel for the spreadsheet version.")
